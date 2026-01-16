import boto3
import pandas as pd
import json
import re
from typing import List, Dict
from pathlib import Path
import sys

def extract_tables_from_image(image_path: str) -> List[pd.DataFrame]:
    """
    Extrae tablas de una imagen o PDF usando Amazon Textract de forma eficiente.

    Args:
        image_path: Ruta a la imagen o PDF local

    Returns:
        Lista de DataFrames, uno por cada tabla encontrada
    """
    # Inicializar cliente de Textract
    textract = boto3.client('textract')

    # Detectar si es PDF
    file_extension = Path(image_path).suffix.lower()

    all_dataframes = []

    if file_extension == '.pdf':
        # Usar PyMuPDF para convertir PDF a imágenes (sin dependencias externas)
        try:
            import fitz  # PyMuPDF
        except ImportError:
            print("\nERROR: La libreria 'PyMuPDF' no esta instalada.")
            print("Por favor instala: pip install PyMuPDF")
            sys.exit(1)

        # Convertir PDF a imágenes (una por página)
        print("Convirtiendo PDF a imagenes...")
        try:
            pdf_document = fitz.open(image_path)
            print(f"Se encontraron {len(pdf_document)} pagina(s)")

            # Procesar cada página
            for page_num in range(len(pdf_document)):
                print(f"Procesando pagina {page_num + 1}/{len(pdf_document)}...")

                # Obtener la página
                page = pdf_document[page_num]

                # Convertir página a imagen (matriz de píxeles)
                pix = page.get_pixmap(dpi=300)

                # Convertir a bytes PNG
                image_bytes = pix.tobytes("png")

                # Llamar a Textract
                response = textract.analyze_document(
                    Document={'Bytes': image_bytes},
                    FeatureTypes=['TABLES']
                )

                # Extraer tablas de esta página
                page_dataframes = parse_tables(response)
                all_dataframes.extend(page_dataframes)

            pdf_document.close()

        except Exception as e:
            print(f"\nERROR: No se pudo convertir el PDF: {str(e)}")
            sys.exit(1)

    else:
        # Es una imagen normal (JPEG, PNG, etc.)
        with open(image_path, 'rb') as document:
            image_bytes = document.read()

        # Llamar a Textract - solo feature TABLES para minimizar costos
        response = textract.analyze_document(
            Document={'Bytes': image_bytes},
            FeatureTypes=['TABLES']  # Solo tablas, no FORMS ni QUERIES
        )

        # Extraer tablas del response
        all_dataframes = parse_tables(response)

    return all_dataframes


def parse_tables(response: Dict) -> List[pd.DataFrame]:
    """
    Parsea la respuesta de Textract y convierte las tablas a DataFrames.

    Args:
        response: Respuesta de analyze_document

    Returns:
        Lista de DataFrames
    """
    blocks = response['Blocks']

    # Crear diccionario de bloques por ID para acceso rápido
    block_map = {block['Id']: block for block in blocks}

    # Encontrar todos los bloques de tipo TABLE
    table_blocks = [block for block in blocks if block['BlockType'] == 'TABLE']

    dataframes = []

    for table in table_blocks:
        # Construir matriz de la tabla
        rows_dict = {}

        if 'Relationships' in table:
            for relationship in table['Relationships']:
                if relationship['Type'] == 'CHILD':
                    for cell_id in relationship['Ids']:
                        cell = block_map[cell_id]
                        if cell['BlockType'] == 'CELL':
                            row_index = cell['RowIndex']
                            col_index = cell['ColumnIndex']

                            # Obtener texto de la celda
                            cell_text = get_cell_text(cell, block_map)

                            # Almacenar en diccionario
                            if row_index not in rows_dict:
                                rows_dict[row_index] = {}
                            rows_dict[row_index][col_index] = cell_text

        # Convertir a DataFrame
        if rows_dict:
            # Ordenar filas y columnas
            sorted_rows = sorted(rows_dict.items())

            # Construir lista de listas
            table_data = []
            for _, row in sorted_rows:
                sorted_cols = sorted(row.items())
                table_data.append([text for _, text in sorted_cols])

            # Crear DataFrame (primera fila como header si existe)
            if len(table_data) > 1:
                df = pd.DataFrame(table_data[1:], columns=table_data[0])
            else:
                df = pd.DataFrame(table_data)

            dataframes.append(df)

    return dataframes


def get_cell_text(cell: Dict, block_map: Dict) -> str:
    """
    Extrae el texto de una celda.

    Args:
        cell: Bloque de celda
        block_map: Diccionario de bloques por ID

    Returns:
        Texto de la celda
    """
    text = ""
    if 'Relationships' in cell:
        for relationship in cell['Relationships']:
            if relationship['Type'] == 'CHILD':
                for word_id in relationship['Ids']:
                    word = block_map.get(word_id)
                    if word and word['BlockType'] == 'WORD':
                        text += word.get('Text', '') + " "
    return text.strip()


def limpiar_datos(df: pd.DataFrame, tipo_operacion: str = 'entrada') -> pd.DataFrame:
    """
    Redirige a la función correcta según el tipo de operación.

    Args:
        df: DataFrame raw extraído de la imagen
        tipo_operacion: 'entrada' o 'salida'

    Returns:
        DataFrame limpio con columnas 'Producto' y 'Cantidad'
    """
    if tipo_operacion.lower() == 'salida':
        config_path = Path(__file__).parent / 'config.json'
        return limpiar_datos_salida(df, str(config_path))
    else:
        return limpiar_datos_entrada(df)


def limpiar_datos_entrada(df: pd.DataFrame) -> pd.DataFrame:
    """
    Limpia el DataFrame para ENTRADA (pedidos).

    COMPORTAMIENTO:
    1. Busca columna de cantidad (palabras clave: 'cantidad', 'cant', 'qty')
    2. Detecta si primera columna es ID numérico (>70% valores numéricos)
       - Si es numérico: usa segunda columna como producto
       - Si no: usa primera columna como producto
    3. Elimina filas vacías y convierte cantidades a números
    4. Limpia prefijos del producto

    Returns:
        DataFrame con columnas: ['Producto', 'Cantidad']
    """
    df_clean = df.copy()

    # Normalizar nombres de columnas
    df_clean.columns = df_clean.columns.str.strip().str.lower()

    # Buscar columna de cantidad
    cantidad_col = None

    # ENTRADA: Busca columna "Cantidad" explícitamente (pedidos con encabezado)
    print("  * Modo ENTRADA: Buscando columna 'Cantidad'...")
    for col in df_clean.columns:
        col_lower = str(col).lower()
        # IMPORTANTE: Buscar 'cantidad' primero (exacto), no 'unid' que puede confundirse con 'unidad'
        if 'cantidad' in col_lower or 'cant.' in col_lower or 'qty' in col_lower:
            cantidad_col = col
            print(f"  [OK] Columna de cantidad encontrada: '{col}'")
            break

    if cantidad_col is None:
        # Intentar usar la última columna como cantidad
        if len(df_clean.columns) >= 2:
            cantidad_col = df_clean.columns[-1]
            print(f"  [ADVERTENCIA] No se encontro 'Cantidad', usando ultima columna: '{cantidad_col}'")
        else:
            raise ValueError(f"No se encontró columna de Cantidad. Columnas: {list(df_clean.columns)}")

    # Detectar si la primera columna es numérica (ID/Referencia)
    # En ese caso, usar la segunda columna como producto
    if len(df_clean.columns) == 0:
        raise ValueError(f"El DataFrame no tiene columnas después de normalizar. DataFrame original tenía: {list(df.columns)}")

    if len(df_clean) == 0:
        raise ValueError(f"El DataFrame no tiene filas después de cargar. Verifica que la imagen tenga contenido.")

    primera_col = df_clean.columns[0]

    # Intentar convertir la primera columna a numérico y ver cuántos valores son válidos
    # IMPORTANTE: Filtrar filas con valores no vacíos ANTES de calcular el porcentaje
    try:
        primera_col_data = df_clean[primera_col]
        if primera_col_data is None or len(primera_col_data) == 0:
            print(f"  [ADVERTENCIA] Primera columna '{primera_col}' esta vacia")
            porcentaje_numerico = 0
        else:
            # Filtrar valores no vacíos (excluir NaN y strings vacíos)
            valores_no_vacios = primera_col_data.dropna()
            valores_no_vacios = valores_no_vacios[valores_no_vacios.astype(str).str.strip() != '']

            if len(valores_no_vacios) == 0:
                porcentaje_numerico = 0
            else:
                # Convertir a numérico solo los valores no vacíos
                valores_numericos = pd.to_numeric(valores_no_vacios, errors='coerce')
                # Calcular porcentaje sobre valores no vacíos (no sobre todo el DataFrame)
                porcentaje_numerico = valores_numericos.notna().sum() / len(valores_no_vacios)
    except Exception as e:
        print(f"  [ERROR] Error al analizar primera columna: {e}")
        print(f"  [ERROR] Columnas: {list(df_clean.columns)}")
        print(f"  [ERROR] Primeras filas del DataFrame:")
        print(df_clean.head())
        porcentaje_numerico = 0

    # Si más del 70% de los valores en la primera columna son numéricos, usar la segunda columna
    if porcentaje_numerico > 0.7 and len(df_clean.columns) > 1:
        print(f"  * Primera columna '{primera_col}' es numerica ({porcentaje_numerico:.0%}), usando segunda columna como Producto")
        producto_col = df_clean.columns[1]
    else:
        print(f"  * Usando primera columna '{primera_col}' como Producto")
        producto_col = primera_col

    # Seleccionar solo esas columnas y renombrar
    df_clean = df_clean[[producto_col, cantidad_col]].copy()
    df_clean.columns = ['Producto', 'Cantidad']

    # Limpiar valores nulos y vacíos
    df_clean = df_clean.dropna(subset=['Producto', 'Cantidad'])
    # Convertir Producto a string antes de usar .str accessor
    df_clean['Producto'] = df_clean['Producto'].astype(str)
    df_clean = df_clean[df_clean['Producto'].str.strip() != '']

    # Convertir cantidad a numérico (reemplazar comas por puntos primero)
    df_clean['Cantidad'] = df_clean['Cantidad'].astype(str).str.replace(',', '.')
    df_clean['Cantidad'] = pd.to_numeric(df_clean['Cantidad'], errors='coerce')
    df_clean = df_clean.dropna(subset=['Cantidad'])

    # Limpiar espacios en producto
    df_clean['Producto'] = df_clean['Producto'].str.strip()

    # Eliminar prefijos numéricos (1., 1-, 14.-, etc.) del inicio del nombre del producto
    # Patrón: número(s) + punto/guión/espacio al inicio
    df_clean['Producto'] = df_clean['Producto'].apply(
        lambda x: re.sub(r'^\d+[\.\-\s]+[\|\s]*', '', x).strip()
    )

    # Eliminar prefijos de error de OCR (I, |, i) al inicio del producto
    df_clean['Producto'] = df_clean['Producto'].apply(
        lambda x: re.sub(r'^[I\|i]\s*', '', x).strip()
    )

    return df_clean.reset_index(drop=True)


def limpiar_datos_salida(df: pd.DataFrame, config_path: str = 'config.json') -> pd.DataFrame:
    """
    Limpia el DataFrame de salida (ventas) filtrando solo productos válidos.

    COMPORTAMIENTO CRÍTICO:
    1. Carga productos válidos desde config.json campo "salida"
    2. Solo extrae líneas que coincidan con productos válidos
    3. Filtra datos irrelevantes (Beneficio, cajero, totales, etc.)
    4. DETECCIÓN INTELIGENTE DE CANTIDAD:
       - Busca en todas las columnas (segunda, tercera, etc.)
       - CASO 1: Si la celda contiene múltiples números separados por espacios
         Ejemplo: "294.800 30" o "57 182.100"
         → Divide por espacios y busca el número ENTERO
       - CASO 2: Si la celda contiene un solo número
         → Verifica si es entero (cantidad) o decimal (precio)
       - Identifica números ENTEROS como cantidad (67, 27, 30)
       - Ignora números con decimales (precios: 294.800, 59.400)
    5. Usa normalización de texto para comparar productos

    Returns:
        DataFrame con columnas: ['Producto', 'Cantidad']
    """
    # Cargar productos válidos
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    productos_validos_salida = set()
    for categoria, info in config.items():
        if not isinstance(info, dict) or 'variantes' not in info:
            continue
        for variante in info['variantes']:
            salidas = variante.get('salida', [])
            for salida in salidas:
                productos_validos_salida.add(normalizar_texto(salida))

    if not productos_validos_salida:
        raise ValueError("No se encontraron productos de salida válidos en config.json")

    print(f"  * Cargados {len(productos_validos_salida)} productos válidos de salida")

    df_clean = df.copy()
    if len(df_clean.columns) > 0:
        df_clean.columns = df_clean.columns.astype(str).str.strip().str.lower()

    resultados = []
    productos_filtrados = []

    for idx, row in df_clean.iterrows():
        # Convertir toda la fila a lista de strings
        valores = [str(v).strip() if pd.notna(v) else '' for v in row.values]

        if not any(valores):
            continue

        # Saltar encabezados con "plu"
        if valores and 'plu' in valores[0].lower():
            continue

        # Buscar producto en primera columna
        if len(valores) >= 2 and valores[0]:
            producto = valores[0]
            producto_normalizado = normalizar_texto(producto)

            # FILTRO: verificar si es producto válido
            es_producto_valido = False
            for producto_config in productos_validos_salida:
                if producto_normalizado in producto_config or producto_config in producto_normalizado:
                    es_producto_valido = True
                    break

            if not es_producto_valido:
                productos_filtrados.append(producto)
                continue

            # DETECCIÓN INTELIGENTE: Buscar la columna con números enteros
            # La cantidad es un número entero (67, 27)
            # El precio tiene decimales (294.800, 59.400)
            cantidad = None

            # Revisar columnas desde la segunda en adelante
            for i in range(1, len(valores)):
                if not valores[i]:
                    continue

                valor_celda = valores[i]

                # CASO 1: La celda contiene múltiples números separados por espacios
                # Ejemplo: "294.800 30" o "57 182.100"
                if ' ' in valor_celda:
                    # Dividir por espacios y buscar números
                    partes = valor_celda.split()
                    for parte in partes:
                        parte_limpia = parte.replace(',', '.')
                        try:
                            num = float(parte_limpia)
                            # Si es entero, es la cantidad
                            if num > 0 and num == int(num):
                                cantidad = int(num)
                                break
                        except (ValueError, TypeError):
                            continue

                    if cantidad is not None:
                        break

                # CASO 2: La celda contiene un solo número
                else:
                    valor_str = valor_celda.replace(',', '.')
                    try:
                        valor_num = float(valor_str)

                        # Si es un número entero (sin decimales significativos)
                        # Ejemplo: 67.0 == 67, pero 59.4 != 59
                        if valor_num > 0 and valor_num == int(valor_num):
                            cantidad = int(valor_num)
                            break
                    except (ValueError, TypeError):
                        continue

            # Si encontramos una cantidad válida, agregar el resultado
            if cantidad is not None and cantidad > 0:
                resultados.append({
                    'Producto': producto,
                    'Cantidad': cantidad
                })
            else:
                # Debug: mostrar qué valores se encontraron
                print(f"  [ADVERTENCIA] No se encontro cantidad entera para '{producto}': valores = {valores[1:]}")

    if productos_filtrados:
        productos_unicos = list(set(productos_filtrados))[:5]
        print(f"  * Filtrados {len(productos_filtrados)} datos no-inventario: {', '.join(productos_unicos)}...")

    if not resultados:
        raise ValueError("No se encontraron productos válidos en los datos de salida")

    df_final = pd.DataFrame(resultados)
    df_final['Producto'] = df_final['Producto'].str.strip()
    df_final['Producto'] = df_final['Producto'].apply(
        lambda x: re.sub(r'^\d+[\.\-\s]+[\|\s]*', '', x).strip()
    )
    df_final['Producto'] = df_final['Producto'].apply(
        lambda x: re.sub(r'^[I\|i]\s*', '', x).strip()
    )

    print(f"  [OK] Procesados {len(df_final)} productos de salida (ventas)")
    return df_final.reset_index(drop=True)


def normalizar_texto(texto: str) -> str:
    """
    Normaliza un texto eliminando espacios, puntos, guiones y caracteres especiales.
    Mantiene números y letras. Convierte todo a minúsculas para facilitar comparaciones.

    Args:
        texto: Texto a normalizar

    Returns:
        Texto normalizado (solo letras y números en minúsculas, sin espacios ni puntuación)
    """
    # Convertir a minúsculas
    texto = texto.lower().strip()

    # Eliminar puntos, guiones, espacios, y caracteres especiales comunes
    # Mantener solo letras y números
    texto_limpio = re.sub(r'[^a-záéíóúñ0-9]', '', texto)

    return texto_limpio


def validar_y_multiplicar(df_clean: pd.DataFrame, config_path: str = 'config.json', tipo_operacion: str = 'entrada') -> pd.DataFrame:
    """
    Redirige a la función correcta según el tipo de operación.
    """
    if tipo_operacion.lower() == 'salida':
        return validar_y_multiplicar_salida(df_clean, config_path)
    else:
        return validar_y_multiplicar_entrada(df_clean, config_path)


def validar_y_multiplicar_entrada(df_clean: pd.DataFrame, config_path: str = 'config.json') -> pd.DataFrame:
    """
    Valida productos contra config.json y MULTIPLICA las cantidades.

    COMPORTAMIENTO:
    1. Busca cada producto en el campo "entrada" de config.json
    2. Usa normalización de texto para comparar
    3. MULTIPLICA la cantidad por el multiplicador definido
       Ejemplo: 2 cajas × multiplicador 12 = 24 unidades
    4. Productos no encontrados: marca como "(no registrado)" con multiplicador 1

    Returns:
        DataFrame con columnas:
        ['Producto', 'Cantidad_Original', 'Multiplicador', 'Cantidad_Final', 'Categoria']
    """
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    resultados = []
    productos_no_encontrados = []

    for _, row in df_clean.iterrows():
        producto = row['Producto']
        cantidad = row['Cantidad']
        producto_normalizado = normalizar_texto(producto)
        encontrado = False

        for categoria, info in config.items():
            if not isinstance(info, dict) or 'variantes' not in info:
                continue

            variantes = info['variantes']
            for variante in variantes:
                entradas = variante.get('entrada', [])
                multiplicador = variante['multiplicador']

                for entrada in entradas:
                    entrada_normalizada = normalizar_texto(entrada)
                    if entrada_normalizada in producto_normalizado or producto_normalizado in entrada_normalizada:
                        cantidad_final = cantidad * multiplicador
                        resultados.append({
                            'Producto': producto,
                            'Cantidad_Original': cantidad,
                            'Multiplicador': multiplicador,
                            'Cantidad_Final': cantidad_final,
                            'Categoria': categoria
                        })
                        encontrado = True
                        break
                if encontrado:
                    break
            if encontrado:
                break

        if not encontrado:
            productos_no_encontrados.append(producto)
            categoria_no_registrada = f"{producto} (no registrado)"
            print(f"  [ADVERTENCIA] Producto no registrado: '{producto}'")
            resultados.append({
                'Producto': producto,
                'Cantidad_Original': cantidad,
                'Multiplicador': 1,
                'Cantidad_Final': cantidad,
                'Categoria': categoria_no_registrada
            })

    if productos_no_encontrados:
        print(f"\n[ADVERTENCIA] Se encontraron {len(productos_no_encontrados)} producto(s) no registrado(s) en config.json")

    return pd.DataFrame(resultados)


def validar_y_multiplicar_salida(df_clean: pd.DataFrame, config_path: str = 'config.json') -> pd.DataFrame:
    """
    Valida productos de salida contra config.json SIN multiplicar cantidades.

    COMPORTAMIENTO:
    1. Busca cada producto en el campo "salida" de config.json
    2. Usa normalización de texto para comparar
    3. NO MULTIPLICA la cantidad (siempre multiplicador = 1)
       Las ventas ya vienen en unidades individuales
    4. Productos no encontrados: marca como "(no registrado)"

    Returns:
        DataFrame con columnas:
        ['Producto', 'Cantidad_Original', 'Multiplicador', 'Cantidad_Final', 'Categoria']
        Nota: Multiplicador siempre es 1, Cantidad_Final = Cantidad_Original
    """
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    resultados = []
    productos_no_encontrados = []

    for _, row in df_clean.iterrows():
        producto = row['Producto']
        cantidad = row['Cantidad']
        producto_normalizado = normalizar_texto(producto)
        encontrado = False

        for categoria, info in config.items():
            if not isinstance(info, dict) or 'variantes' not in info:
                continue

            variantes = info['variantes']
            for variante in variantes:
                salidas = variante.get('salida', [])
                if not salidas:
                    continue

                for salida in salidas:
                    salida_normalizada = normalizar_texto(salida)
                    if salida_normalizada in producto_normalizado or producto_normalizado in salida_normalizada:
                        resultados.append({
                            'Producto': producto,  # Mantener el nombre original de la factura
                            'Cantidad_Original': cantidad,
                            'Multiplicador': 1,
                            'Cantidad_Final': cantidad,
                            'Categoria': categoria
                        })
                        encontrado = True
                        break
                if encontrado:
                    break
            if encontrado:
                break

        if not encontrado:
            productos_no_encontrados.append(producto)
            categoria_no_registrada = f"{producto} (no registrado)"
            print(f"  [ADVERTENCIA] Producto de salida no registrado: '{producto}'")
            resultados.append({
                'Producto': producto,
                'Cantidad_Original': cantidad,
                'Multiplicador': 1,
                'Cantidad_Final': cantidad,
                'Categoria': categoria_no_registrada
            })

    if productos_no_encontrados:
        print(f"\n[ADVERTENCIA] Se encontraron {len(productos_no_encontrados)} producto(s) de salida no registrado(s)")

    return pd.DataFrame(resultados)


def actualizar_inventario_layout(df_final: pd.DataFrame, layout_path: str = 'Inventario_layout.xlsx', tipo_operacion: str = 'entrada', output_path: str = None) -> str:
    """
    Actualiza el archivo de inventario con las cantidades finales por categoría.

    Preserva TODO el estilo del Excel: bordes, colores, fuentes, espaciado, etc.
    Busca la columna especificada (entrada o salida) en los encabezados y la fila con el nombre de la categoría,
    luego coloca el valor en la intersección.

    IMPORTANTE - Lógica de fechas (compara nombre del archivo VS fecha del calendario):

    1. MISMA FECHA (continuación de carga):
       - Usuario carga: "inventario_12_12_2025.xlsx"
       - Usuario selecciona en calendario: 12/12/2025
       - Resultado: Las fechas coinciden → CONTINUACIÓN de carga
       - BASE: El archivo que cargaste (inventario_12_12_2025.xlsx)
       - Comportamiento:
         * NO copia Inv Final → Inv Inicial (mantiene todo intacto)
         * Agrega/sobrescribe valores en columna "Entrada" o "Salida"
         * Guarda como: "inventario_12_12_2025.xlsx" en Descargas (sobrescribe)

    2. FECHA DIFERENTE (nuevo día):
       - Usuario carga: "inventario_12_12_2025.xlsx" (del día anterior)
       - Usuario selecciona en calendario: 13/12/2025 (nuevo día)
       - Resultado: Las fechas NO coinciden → NUEVO DÍA
       - BASE: "Inventario_layout.xlsx" (template de .sistema)
       - Comportamiento:
         * Copia Inv Final (col E) del archivo cargado → Inv Inicial (col B) del template
         * Agrega valores en columna "Entrada" o "Salida"
         * Guarda como: "inventario_13_12_2025.xlsx" en Descargas

    3. PRIMER ARCHIVO:
       - Usuario carga: "Inventario_layout.xlsx" (template vacío de .sistema)
       - Usuario selecciona en calendario: 10/12/2025
       - BASE: "Inventario_layout.xlsx" (el que cargó)
       - Comportamiento:
         * NO hay Inv Final anterior para copiar
         * Agrega valores en columna "Entrada" o "Salida"
         * Guarda como: "inventario_10_12_2025.xlsx" en Descargas

    Args:
        df_final: DataFrame con las cantidades finales por categoría
        layout_path: Archivo que el usuario sube (inventario anterior o template)
        tipo_operacion: 'entrada' o 'salida' - determina qué columna usar en el Excel
        output_path: Ruta donde guardar el archivo actualizado (nombre generado con fecha del calendario)

    Returns:
        Ruta del archivo guardado
    """
    from openpyxl import load_workbook
    from copy import copy
    import os
    import re

    try:
        # PASO 1: Comparar fecha del archivo cargado VS fecha seleccionada en el calendario
        misma_fecha = False

        # Extraer fecha del nombre del archivo que el usuario carga
        # Ejemplo: "inventario_12_12_2025.xlsx" → fecha_archivo = "12_12_2025"
        layout_filename = Path(layout_path).name
        match_archivo = re.search(r'inventario_(\d{2}_\d{2}_\d{4})\.xlsx', layout_filename)

        # Extraer fecha del output_path (que se genera con la fecha del calendario)
        # Ejemplo: "C:\Users\...\Downloads\inventario_13_12_2025.xlsx" → fecha_calendario = "13_12_2025"
        if output_path and match_archivo:
            output_filename = Path(output_path).name
            match_calendario = re.search(r'inventario_(\d{2}_\d{2}_\d{4})\.xlsx', output_filename)

            if match_calendario:
                fecha_archivo = match_archivo.group(1)
                fecha_calendario = match_calendario.group(1)

                if fecha_archivo == fecha_calendario:
                    # MISMA FECHA: inventario_12_12_2025.xlsx + calendario 12/12/2025
                    misma_fecha = True
                    print(f"  [DETECTADO] MISMA FECHA")
                    print(f"  - Archivo cargado: inventario_{fecha_archivo}.xlsx")
                    print(f"  - Fecha seleccionada: {fecha_calendario}")
                    print(f"  - Se copiara TODO el archivo tal cual (continuacion de carga)")
                    print(f"  - Inv Inicial e Inv Final NO se modificaran")
                else:
                    # FECHA DIFERENTE: inventario_12_12_2025.xlsx + calendario 13/12/2025
                    misma_fecha = False
                    print(f"  [OK] FECHA DIFERENTE")
                    print(f"  - Archivo cargado (dia anterior): inventario_{fecha_archivo}.xlsx")
                    print(f"  - Fecha seleccionada (nuevo dia): {fecha_calendario}")
                    print(f"  - Se copiara Inv Final (col E) a Inv Inicial (col B)")

        # PASO 2: Determinar qué archivo usar como base
        # IMPORTANTE: Si el archivo de salida YA EXISTE, usarlo como base para preservar
        # todas las columnas (Entrada, Salida, etc.)
        if output_path and os.path.exists(output_path):
            # El archivo de salida ya existe, usarlo como base para preservar datos
            archivo_base = output_path
            archivo_para_inv_final = None
            misma_fecha = True  # Forzar misma_fecha porque estamos continuando
            print(f"  - Base: archivo de salida existente (preservar todas las columnas)")
            print(f"  - Archivo: {Path(output_path).name}")
        elif misma_fecha:
            # MISMA FECHA: usar el archivo que el usuario cargó
            archivo_base = layout_path
            archivo_para_inv_final = None  # No necesitamos copiar Inv Final
            print(f"  - Base: archivo que cargaste (misma fecha)")
        else:
            # FECHA DIFERENTE: usar el template de .sistema
            template_path = Path(__file__).parent / 'Inventario_layout.xlsx'
            archivo_base = str(template_path)
            archivo_para_inv_final = layout_path  # Copiar Inv Final de este archivo
            print(f"  - Base: Inventario_layout.xlsx (template de .sistema)")
            print(f"  - Inv Final se copiara desde: {Path(layout_path).name}")

        # Cargar el archivo base
        wb_original = load_workbook(archivo_base, data_only=False, keep_vba=False)
        ws_original = wb_original.active

        # SIEMPRE cargar valores del archivo base (template o archivo mismo)
        wb_valores = load_workbook(archivo_base, data_only=True, keep_vba=False)
        ws_valores = wb_valores.active

        # Si es FECHA DIFERENTE, cargar ADICIONAL el archivo anterior para copiar Inv Final
        # IMPORTANTE: NO usar data_only=True porque no funciona si Excel no guardó valores.
        # En su lugar, cargaremos con data_only=False y CALCULAREMOS manualmente Inv Final
        wb_inv_final_anterior = None
        ws_inv_final_anterior = None
        if archivo_para_inv_final:
            # Cargar el archivo con data_only=False para leer TODO (incluyendo fórmulas)
            wb_inv_final_anterior = load_workbook(archivo_para_inv_final, data_only=False, keep_vba=False)
            ws_inv_final_anterior = wb_inv_final_anterior.active

            # TAMBIÉN cargar con data_only=True para INTENTAR leer valores calculados
            # (en caso de que Excel sí haya guardado los valores)
            wb_inv_final_valores = load_workbook(archivo_para_inv_final, data_only=True, keep_vba=False)
            ws_inv_final_valores = wb_inv_final_valores.active

        # Crear workbook nuevo para resultado final
        wb = load_workbook(archivo_base, data_only=False, keep_vba=False)
        ws = wb.active

        # PASO CRÍTICO: Copiar COLUMNA E → COLUMNA B solo si es FECHA DIFERENTE
        COL_INV_FINAL = 5   # Columna E (5 en índice de Excel)
        COL_INV_INICIAL = 2  # Columna B (2 en índice de Excel)

        if not misma_fecha and ws_inv_final_anterior:
            # Copiar COLUMNA E (Inv Final) del archivo ORIGEN → COLUMNA B (Inv Inicial) del archivo NUEVO
            # Columna E = Inv Final del archivo anterior (origen)
            # Columna B = Inv Inicial del archivo nuevo (destino)
            print("  - Copiando Columna E (Inv Final del archivo anterior) a Columna B (Inv Inicial del archivo nuevo)...")

            # COLUMNAS DE LA FÓRMULA: Inv Final = B (Inv Inicial) + C (Entrada) - D (Salida)
            COL_B_INV_INICIAL = 2
            COL_C_ENTRADA = 3
            COL_D_SALIDA = 4

            valores_copiados = 0
            valores_nulos = 0
            valores_calculados_manual = 0

            for row_idx in range(2, ws_inv_final_anterior.max_row + 1):  # Empezar desde fila 2 (saltar encabezado)
                # ESTRATEGIA 1: Intentar leer el valor calculado de columna E (si Excel lo guardó)
                cell_inv_final_valores = ws_inv_final_valores.cell(row=row_idx, column=COL_INV_FINAL)
                valor_final = cell_inv_final_valores.value

                # Si data_only=True devolvió un valor numérico válido, úsalo
                if valor_final is not None and isinstance(valor_final, (int, float)):
                    # Tenemos el valor calculado, usarlo directamente
                    pass  # valor_final ya está asignado
                else:
                    # ESTRATEGIA 2: data_only=True no funcionó, CALCULAR MANUALMENTE
                    # La fórmula de Inv Final es: =B + C - D
                    # Leer valores de columnas B, C, D del archivo anterior
                    val_b = ws_inv_final_valores.cell(row=row_idx, column=COL_B_INV_INICIAL).value
                    val_c = ws_inv_final_valores.cell(row=row_idx, column=COL_C_ENTRADA).value
                    val_d = ws_inv_final_valores.cell(row=row_idx, column=COL_D_SALIDA).value

                    # Convertir None a 0 para el cálculo
                    val_b = val_b if val_b is not None and isinstance(val_b, (int, float)) else 0
                    val_c = val_c if val_c is not None and isinstance(val_c, (int, float)) else 0
                    val_d = val_d if val_d is not None and isinstance(val_d, (int, float)) else 0

                    # CALCULAR: Inv Final = Inv Inicial + Entrada - Salida
                    valor_final = val_b + val_c - val_d
                    valores_calculados_manual += 1

                # VALIDACIÓN: Solo copiar si es un NÚMERO válido (int o float)
                if valor_final is not None and isinstance(valor_final, (int, float)):
                    # Copiar el VALOR NUMÉRICO a columna B (Inv Inicial) del archivo NUEVO
                    cell_inv_inicial_destino = ws.cell(row=row_idx, column=COL_INV_INICIAL)

                    # CRÍTICO: Asignar como NÚMERO (float o int), NUNCA como fórmula
                    # Si valor_final es int, mantenerlo como int; si es float, mantenerlo como float
                    if isinstance(valor_final, int):
                        cell_inv_inicial_destino.value = int(valor_final)
                    else:
                        cell_inv_inicial_destino.value = float(valor_final)

                    valores_copiados += 1
                elif valor_final is None:
                    # Celda vacía o fórmula sin valor calculado
                    valores_nulos += 1

            print(f"  [OK] {valores_copiados} valores numéricos copiados: Columna E (Inv Final anterior) → Columna B (Inv Inicial nuevo)")
            if valores_calculados_manual > 0:
                print(f"  [INFO] {valores_calculados_manual} valores calculados manualmente (B+C-D)")
            if valores_nulos > 0:
                print(f"  [INFO] {valores_nulos} celdas vacías o sin valor calculado fueron omitidas")

            # Cerrar el workbook auxiliar de valores
            wb_inv_final_valores.close()
        else:
            # MISMA FECHA: NO copiar, mantener el Inv Inicial actual
            print("  [OK] Inv Inicial NO modificado (misma fecha de inventario, continuacion de carga)")

        # Identificar columnas que NO deben ser sobrescritas durante la conversión
        print("  - Identificando columnas a preservar...")
        columnas_a_no_sobrescribir = set()

        # Columna B (2) = Inv Inicial - NO sobrescribir (acabamos de copiar valores aquí)
        # Columna E (5) = Inv Final - DEBE mantener fórmulas
        # Columna G (7) - DEBE mantener fórmulas
        columnas_a_no_sobrescribir.add(2)  # Columna B (Inv Inicial) - RECIÉN COPIADA
        columnas_a_no_sobrescribir.add(5)  # Columna E (Inv Final)
        columnas_a_no_sobrescribir.add(7)  # Columna G

        # IMPORTANTE: Preservar TODAS las columnas de operación (Entrada, Salida, etc.)
        # excepto la que estamos modificando en esta carga
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            cell_value = col[0].value
            if cell_value:
                cell_lower = str(cell_value).lower().strip()
                # Si es una columna de operación DIFERENTE a la que estamos cargando, protegerla
                if cell_lower in ['entrada', 'salida'] and cell_lower != tipo_operacion.lower():
                    columnas_a_no_sobrescribir.add(col_idx)
                    print(f"  [OK] Columna '{cell_value}' NO sera sobrescrita (preservar datos existentes)")

        print(f"  [OK] Columna B (Inv Inicial) NO sera sobrescrita (valores recien copiados)")
        print(f"  [OK] Columna E (Inv Final) mantendra formulas")
        print(f"  [OK] Columna G mantendra formulas")

        # Convertir fórmulas a valores EXCEPTO en las columnas protegidas
        print("  - Convirtiendo formulas a valores (excepto columnas protegidas)...")
        formulas_convertidas = 0
        formulas_preservadas = 0

        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                # SALTAR columnas protegidas (B, E, G)
                if col_idx in columnas_a_no_sobrescribir:
                    # No tocar esta celda, ya tiene el valor correcto
                    continue

                cell = ws.cell(row=row_idx, column=col_idx)
                cell_valor = ws_valores.cell(row=row_idx, column=col_idx)

                # Si la celda tiene una fórmula
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    # Convertir a valor
                    if cell_valor.value is not None:
                        cell.value = cell_valor.value
                        formulas_convertidas += 1
                    else:
                        # Si no hay valor, poner 0
                        cell.value = 0
                        formulas_convertidas += 1
                elif cell_valor.value is not None and cell.value != cell_valor.value:
                    # Asegurar que usamos valores, no fórmulas
                    cell.value = cell_valor.value

        print(f"  [OK] {formulas_convertidas} formulas convertidas a valores")
        print(f"  [OK] Columnas B, E, G preservadas correctamente")

        # Cerrar workbooks auxiliares
        wb_original.close()
        wb_valores.close()
        if wb_inv_final_anterior:
            wb_inv_final_anterior.close()

        # Buscar la columna que contiene el tipo de operación en la primera fila (encabezados)
        col_entrada_idx = None
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            cell_value = col[0].value
            if cell_value and str(cell_value).lower().strip() == tipo_operacion.lower():
                col_entrada_idx = col_idx
                break

        if col_entrada_idx is None:
            print(f"  [ERROR] No se encontro la columna '{tipo_operacion}' en {layout_path}")
            return

        # PASO 1: Leer cantidades EXISTENTES del archivo de SALIDA (no del template)
        # para hacer merge inteligente
        print("  - Leyendo cantidades existentes del archivo de salida...")
        cantidades_existentes = {}

        # Verificar si el archivo de salida ya existe
        if output_path and os.path.exists(output_path):
            print(f"    Archivo de salida ya existe: {Path(output_path).name}")
            # Cargar el archivo de salida existente para leer sus valores
            wb_salida_existente = load_workbook(output_path, data_only=True, keep_vba=False)
            ws_salida_existente = wb_salida_existente.active

            for fila_idx, row in enumerate(ws_salida_existente.iter_rows(min_col=1, max_col=1), start=1):
                cell_value = row[0].value
                if cell_value and str(cell_value).strip():
                    categoria = str(cell_value).strip()
                    # Leer el valor actual de la columna de entrada/salida
                    valor_actual = ws_salida_existente.cell(row=fila_idx, column=col_entrada_idx).value
                    if valor_actual is not None and isinstance(valor_actual, (int, float)) and valor_actual > 0:
                        cantidades_existentes[categoria] = valor_actual
                        print(f"    Existente: '{categoria}' = {valor_actual}")

            wb_salida_existente.close()
        else:
            print(f"    Archivo de salida no existe aún, primera carga para esta fecha")

        # PASO 2: Agrupar cantidades NUEVAS por categoría
        cantidades_nuevas = df_final.groupby('Categoria')['Cantidad_Final'].sum()

        # PASO 3: MERGE INTELIGENTE - combinar existentes + nuevas
        print("  - Haciendo merge inteligente de cantidades...")
        cantidades_por_categoria = {}

        # Agregar todas las cantidades existentes primero
        for categoria, cantidad in cantidades_existentes.items():
            cantidades_por_categoria[categoria] = cantidad

        # Sobrescribir SOLO las categorías que vienen con datos nuevos (> 0)
        for categoria, cantidad_nueva in cantidades_nuevas.items():
            if cantidad_nueva and cantidad_nueva > 0:
                # Sobrescribir con el nuevo valor
                cantidades_por_categoria[categoria] = cantidad_nueva
                if categoria in cantidades_existentes:
                    print(f"    Merge: '{categoria}': {cantidades_existentes[categoria]} → {cantidad_nueva} (actualizado)")
                else:
                    print(f"    Merge: '{categoria}': (nuevo) → {cantidad_nueva}")
            else:
                # Cantidad nueva es 0 o None, preservar existente
                if categoria in cantidades_existentes:
                    print(f"    Merge: '{categoria}': {cantidades_existentes[categoria]} (preservado, nueva carga trae {cantidad_nueva})")
                # Si no existe y la nueva es 0, no hacer nada (no agregamos 0s)

        # Actualizar o crear filas para cada categoría (ahora con merge ya hecho)
        print("  - Actualizando celdas en el Excel...")
        for categoria, cantidad in cantidades_por_categoria.items():
            if categoria == 'Sin Categoria':
                continue  # Ignorar productos sin categoría

            # Buscar la fila que contiene el nombre de la categoría en la primera columna
            fila_encontrada = False
            for fila_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=1), start=1):
                cell_value = row[0].value
                if cell_value and str(cell_value).strip() == categoria:
                    # Actualizar la celda en la intersección
                    target_cell = ws.cell(row=fila_idx, column=col_entrada_idx)
                    target_cell.value = cantidad
                    print(f"  + Celda actualizada '{categoria}': {cantidad}")
                    fila_encontrada = True
                    break

            if not fila_encontrada:
                # Encontrar la última fila con datos reales en la primera columna
                ultima_fila_real = 1
                for fila_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=fila_idx, column=1).value
                    if cell_value and str(cell_value).strip():
                        ultima_fila_real = fila_idx

                # Crear nueva fila inmediatamente después del último producto
                nueva_fila = ultima_fila_real + 1

                # Copiar estilos de la fila anterior si existe
                if ultima_fila_real > 1:
                    for col_idx in range(1, ws.max_column + 1):
                        celda_anterior = ws.cell(row=ultima_fila_real, column=col_idx)
                        celda_nueva = ws.cell(row=nueva_fila, column=col_idx)

                        # Copiar estilos (bordes, fuente, relleno, alineación)
                        if celda_anterior.has_style:
                            celda_nueva.font = copy(celda_anterior.font)
                            celda_nueva.border = copy(celda_anterior.border)
                            celda_nueva.fill = copy(celda_anterior.fill)
                            celda_nueva.number_format = copy(celda_anterior.number_format)
                            celda_nueva.protection = copy(celda_anterior.protection)
                            celda_nueva.alignment = copy(celda_anterior.alignment)

                # Asignar valores
                ws.cell(row=nueva_fila, column=1).value = categoria
                ws.cell(row=nueva_fila, column=col_entrada_idx).value = cantidad
                print(f"  + Creada nueva categoria '{categoria}': {cantidad}")

        # Guardar el workbook preservando todos los estilos
        save_path = output_path if output_path else layout_path
        wb.save(save_path)
        wb.close()
        print(f"\n+ Inventario guardado: '{save_path}'")

        if misma_fecha:
            print("  [OK] MISMA FECHA: Se agregaron movimientos al inventario existente")
            print("  [OK] Inv Inicial NO modificado (continuacion de carga de la misma fecha)")
        else:
            print("  [OK] FECHA DIFERENTE: Columna E (Inv Final anterior) a Columna B (Inv Inicial nuevo)")

        print("  [OK] Se preservaron todos los bordes, colores y estilos del Excel")
        print("  [OK] Columna E (Inv Final) mantiene formulas para recalcular automaticamente")

        return save_path

    except FileNotFoundError:
        print(f"\n[ADVERTENCIA] No se encontro el archivo '{layout_path}'")
        return None

    except Exception as e:
        print(f"\nX Error al actualizar inventario: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    # ==================== CONFIGURACIÓN ====================
    # Cargar configuración desde config.json
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)

    USAR_AWS = config.get('USAR_AWS', False)
    #image_path = "WhatsApp Image 2025-11-04 at 10.35.29 PM (1).jpeg"
    image_path = r"C:\Users\Juan Jose\Downloads\ENERO 6.pdf"
    csv_path = "datos_raw.csv"
    # =======================================================

    try:
        if USAR_AWS:
            # PASO 1A: Extraer tablas de la imagen con AWS
            print(f"Analizando imagen: {image_path}")
            print("Extrayendo tablas con Amazon Textract...")

            dataframes = extract_tables_from_image(image_path)

            if not dataframes:
                print("No se encontraron tablas en la imagen")
                exit(1)

            print(f"\nSe encontraron {len(dataframes)} tabla(s)")

            # DEBUG: Mostrar información de todas las tablas detectadas
            if len(dataframes) > 0:
                print("\n[DEBUG] Tablas detectadas por AWS Textract:")
                for idx, df in enumerate(dataframes):
                    print(f"  Tabla {idx + 1}: {df.shape[0]} filas x {df.shape[1]} columnas")
                    print(f"    Columnas: {list(df.columns)}")
                    if len(df) > 0:
                        print(f"    Primera fila: {list(df.iloc[0].values)}")
                    print()

            # Si hay múltiples tablas, filtrar y seleccionar la correcta
            if len(dataframes) > 1:
                print("Filtrando tablas (ignorando encabezados y resúmenes financieros)...")

                # PASO 1: Buscar tablas con columnas de productos
                tablas_con_productos = []
                for df in dataframes:
                    columnas_str = ' '.join([str(col).lower() for col in df.columns])
                    palabras_productos = ['cantidad', 'descripcion', 'descripción', 'referencia', 'producto', 'unidad']

                    tiene_columna_producto = False
                    for palabra in palabras_productos:
                        if palabra in columnas_str:
                            tiene_columna_producto = True
                            break

                    if tiene_columna_producto:
                        tablas_con_productos.append(df)
                        print(f"  + Tabla con columnas de productos detectada ({df.shape[0]} filas x {df.shape[1]} columnas)")

                # PASO 2: Filtrar solo resúmenes financieros (que NO tengan columnas de productos)
                tablas_no_resumen = []
                for df in dataframes:
                    columnas_str = ' '.join([str(col).lower() for col in df.columns])
                    todas_filas_str = ' '.join([str(val).lower() for row in df.values for val in row if pd.notna(val)])

                    # Verificar si tiene columnas de productos
                    palabras_productos = ['cantidad', 'descripcion', 'descripción', 'referencia', 'producto', 'unidad']
                    tiene_columna_producto = any(palabra in columnas_str for palabra in palabras_productos)

                    # Solo es resumen si NO tiene columnas de productos
                    palabras_resumen = ['sub total', 'subtotal', 'total factura', 'total a pagar']
                    es_solo_resumen = (not tiene_columna_producto) and any(palabra in columnas_str or palabra in todas_filas_str for palabra in palabras_resumen)

                    if es_solo_resumen:
                        print(f"  - Descartada tabla de resumen financiero ({df.shape[0]} filas)")
                    else:
                        tablas_no_resumen.append(df)

                # PASO 3: Seleccionar la mejor tabla (priorizar tablas con productos)
                if tablas_con_productos:
                    # Usar la tabla más grande entre las que tienen columnas de productos
                    df_raw = max(tablas_con_productos, key=lambda df: len(df))
                    print(f"Tabla seleccionada: {df_raw.shape[0]} filas x {df_raw.shape[1]} columnas (con columnas de productos)")
                elif tablas_no_resumen:
                    # Si no hay tablas con columnas de productos, usar la más grande que no sea resumen
                    df_raw = max(tablas_no_resumen, key=lambda df: len(df))
                    print(f"Tabla seleccionada: {df_raw.shape[0]} filas x {df_raw.shape[1]} columnas (sin columnas de productos)")
                else:
                    # Último recurso: usar la tabla más grande disponible
                    df_raw = max(dataframes, key=lambda df: len(df))
                    print(f"Tabla seleccionada (sin filtro): {df_raw.shape[0]} filas x {df_raw.shape[1]} columnas")
            else:
                # Solo hay 1 tabla, verificar si es un resumen financiero
                df_raw = dataframes[0]

                # Verificar si es tabla de resumen financiero
                todas_columnas_str = ' '.join([str(col).lower() for col in df_raw.columns])
                todas_filas_str = ' '.join([str(val).lower() for row in df_raw.values for val in row if pd.notna(val)])

                palabras_resumen = ['sub total', 'subtotal', 'descuento', 'iva', 'ibua', 'vr. total', 'total factura']
                es_resumen = False
                for palabra in palabras_resumen:
                    if palabra in todas_columnas_str or palabra in todas_filas_str:
                        es_resumen = True
                        break

                if es_resumen:
                    print("\n[ERROR CRÍTICO] AWS Textract solo detectó la tabla de resumen financiero, no la tabla de productos.")
                    print("Esto puede deberse a:")
                    print("  1. El PDF tiene un formato complejo que Textract no puede interpretar correctamente")
                    print("  2. La tabla de productos está en un formato no estándar (imagen, texto superpuesto, etc.)")
                    print("\nPor favor, intenta:")
                    print("  - Usar un PDF diferente o más simple")
                    print("  - Convertir el PDF a imagen (PNG/JPG) antes de procesar")
                    print("  - Verificar que el PDF no esté protegido o cifrado")
                    sys.exit(1)

            # Guardar CSV para reutilización
            df_raw.to_csv('datos_raw.csv', index=False, encoding='utf-8-sig')
            print("DataFrame guardado en 'datos_raw.csv'")
        else:
            # PASO 1B: Cargar desde CSV
            print(f"Cargando datos desde: {csv_path}")
            df_raw = pd.read_csv(csv_path, encoding='utf-8-sig', thousands=None, decimal='.')
            print("Datos cargados exitosamente")

        # Mostrar datos raw
        print(f"\nDimensiones: {df_raw.shape[0]} filas x {df_raw.shape[1]} columnas")
        print("\nVista previa del DataFrame raw:")
        print(df_raw.head(10).to_string(index=False))

        # PASO 2: Limpiar datos (solo producto y cantidad)
        print("\n" + "="*60)
        print("PASO 2: Limpiando datos...")
        df_clean = limpiar_datos(df_raw)
        print(f"Datos limpios: {len(df_clean)} productos encontrados")
        print("\nDataFrame limpio (Producto y Cantidad):")
        print(df_clean.to_string(index=False))

        # PASO 3: Validar contra config.json y multiplicar
        print("\n" + "="*60)
        print("PASO 3: Validando productos y aplicando multiplicador...")
        df_final = validar_y_multiplicar(df_clean, 'config.json')

        if df_final.empty:
            print("\nADVERTENCIA: No se encontraron productos que coincidan con config.json")
            print("\nProductos en el pedido:")
            for producto in df_clean['Producto'].unique():
                print(f"  - {producto}")
        else:
            print(f"\nProductos validados: {len(df_final)}")
            print("\nDataFrame final:")
            print(df_final.to_string(index=False))

        # PASO 4: Exportar a Excel (siempre exportar, incluso si está vacío)
        print("\n" + "="*60)
        print("PASO 4: Exportando a Excel...")
        output_file = 'productos_final.xlsx'
        df_final.to_excel(output_file, index=False, engine='openpyxl')
        print(f"\nArchivo exportado exitosamente: '{output_file}'")

        # PASO 5: Actualizar Inventario_layout.xlsx
        if not df_final.empty:
            print("\n" + "="*60)
            print("PASO 5: Actualizando Inventario_layout.xlsx...")
            # Por defecto usa 'entrada' en el script standalone
            actualizar_inventario_layout(df_final, 'Inventario_layout.xlsx', tipo_operacion='entrada')

            # Resumen
            print("\n" + "="*60)
            print("RESUMEN:")
            print(f"  - Total productos validados: {len(df_final)}")
            print(f"  - Cantidad total original: {df_final['Cantidad_Original'].sum():.0f}")
            print(f"  - Cantidad total final: {df_final['Cantidad_Final'].sum():.0f}")

    except FileNotFoundError as e:
        print(f"Error: No se encontro el archivo - {str(e)}")
    except ValueError as e:
        print(f"Error de validacion: {str(e)}")
    except Exception as e:
        print(f"Error al procesar la imagen: {str(e)}")
        import traceback
        traceback.print_exc()
