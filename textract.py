import boto3
import pandas as pd
import json
import re
from typing import List, Dict
from pathlib import Path
import sys

def extract_tables_from_image(image_path: str) -> List[pd.DataFrame]:
    """
    Extrae tablas de una imagen o PDF usando Amazon Textract de forma eficiente.

    Args:
        image_path: Ruta a la imagen o PDF local

    Returns:
        Lista de DataFrames, uno por cada tabla encontrada
    """
    # Inicializar cliente de Textract
    textract = boto3.client('textract')

    # Detectar si es PDF
    file_extension = Path(image_path).suffix.lower()

    all_dataframes = []

    if file_extension == '.pdf':
        # Usar PyMuPDF para convertir PDF a imágenes (sin dependencias externas)
        try:
            import fitz  # PyMuPDF
        except ImportError:
            print("\nERROR: La libreria 'PyMuPDF' no esta instalada.")
            print("Por favor instala: pip install PyMuPDF")
            sys.exit(1)

        # Convertir PDF a imágenes (una por página)
        print("Convirtiendo PDF a imagenes...")
        try:
            pdf_document = fitz.open(image_path)
            print(f"Se encontraron {len(pdf_document)} pagina(s)")

            # Procesar cada página
            for page_num in range(len(pdf_document)):
                print(f"Procesando pagina {page_num + 1}/{len(pdf_document)}...")

                # Obtener la página
                page = pdf_document[page_num]

                # Convertir página a imagen (matriz de píxeles)
                pix = page.get_pixmap(dpi=300)

                # Convertir a bytes PNG
                image_bytes = pix.tobytes("png")

                # Llamar a Textract
                response = textract.analyze_document(
                    Document={'Bytes': image_bytes},
                    FeatureTypes=['TABLES']
                )

                # Extraer tablas de esta página
                page_dataframes = parse_tables(response)
                all_dataframes.extend(page_dataframes)

            pdf_document.close()

        except Exception as e:
            print(f"\nERROR: No se pudo convertir el PDF: {str(e)}")
            sys.exit(1)

    else:
        # Es una imagen normal (JPEG, PNG, etc.)
        with open(image_path, 'rb') as document:
            image_bytes = document.read()

        # Llamar a Textract - solo feature TABLES para minimizar costos
        response = textract.analyze_document(
            Document={'Bytes': image_bytes},
            FeatureTypes=['TABLES']  # Solo tablas, no FORMS ni QUERIES
        )

        # Extraer tablas del response
        all_dataframes = parse_tables(response)

    return all_dataframes


def parse_tables(response: Dict) -> List[pd.DataFrame]:
    """
    Parsea la respuesta de Textract y convierte las tablas a DataFrames.

    Args:
        response: Respuesta de analyze_document

    Returns:
        Lista de DataFrames
    """
    blocks = response['Blocks']

    # Crear diccionario de bloques por ID para acceso rápido
    block_map = {block['Id']: block for block in blocks}

    # Encontrar todos los bloques de tipo TABLE
    table_blocks = [block for block in blocks if block['BlockType'] == 'TABLE']

    dataframes = []

    for table in table_blocks:
        # Construir matriz de la tabla
        rows_dict = {}

        if 'Relationships' in table:
            for relationship in table['Relationships']:
                if relationship['Type'] == 'CHILD':
                    for cell_id in relationship['Ids']:
                        cell = block_map[cell_id]
                        if cell['BlockType'] == 'CELL':
                            row_index = cell['RowIndex']
                            col_index = cell['ColumnIndex']

                            # Obtener texto de la celda
                            cell_text = get_cell_text(cell, block_map)

                            # Almacenar en diccionario
                            if row_index not in rows_dict:
                                rows_dict[row_index] = {}
                            rows_dict[row_index][col_index] = cell_text

        # Convertir a DataFrame
        if rows_dict:
            # Ordenar filas y columnas
            sorted_rows = sorted(rows_dict.items())

            # Construir lista de listas
            table_data = []
            for _, row in sorted_rows:
                sorted_cols = sorted(row.items())
                table_data.append([text for _, text in sorted_cols])

            # Crear DataFrame (primera fila como header si existe)
            if len(table_data) > 1:
                df = pd.DataFrame(table_data[1:], columns=table_data[0])
            else:
                df = pd.DataFrame(table_data)

            dataframes.append(df)

    return dataframes


def get_cell_text(cell: Dict, block_map: Dict) -> str:
    """
    Extrae el texto de una celda.

    Args:
        cell: Bloque de celda
        block_map: Diccionario de bloques por ID

    Returns:
        Texto de la celda
    """
    text = ""
    if 'Relationships' in cell:
        for relationship in cell['Relationships']:
            if relationship['Type'] == 'CHILD':
                for word_id in relationship['Ids']:
                    word = block_map.get(word_id)
                    if word and word['BlockType'] == 'WORD':
                        text += word.get('Text', '') + " "
    return text.strip()


def limpiar_datos(df: pd.DataFrame) -> pd.DataFrame:
    """
    Limpia el DataFrame dejando solo columnas de Producto y Cantidad.

    Args:
        df: DataFrame raw extraído de la imagen

    Returns:
        DataFrame limpio con columnas 'Producto' y 'Cantidad'
    """
    df_clean = df.copy()

    # Normalizar nombres de columnas
    df_clean.columns = df_clean.columns.str.strip().str.lower()

    # Buscar columna de cantidad
    cantidad_col = None

    for col in df_clean.columns:
        if 'cantidad' in col or 'cant' in col or 'qty' in col:
            cantidad_col = col
            break

    if cantidad_col is None:
        raise ValueError("No se encontró columna de Cantidad en el DataFrame")

    # Detectar si la primera columna es numérica (ID/Referencia)
    # En ese caso, usar la segunda columna como producto
    primera_col = df_clean.columns[0]

    # Intentar convertir la primera columna a numérico y ver cuántos valores son válidos
    valores_numericos = pd.to_numeric(df_clean[primera_col], errors='coerce')
    porcentaje_numerico = valores_numericos.notna().sum() / len(df_clean)

    # Si más del 70% de los valores en la primera columna son numéricos, usar la segunda columna
    if porcentaje_numerico > 0.7 and len(df_clean.columns) > 1:
        print(f"  * Primera columna '{primera_col}' es numerica ({porcentaje_numerico:.0%}), usando segunda columna como Producto")
        producto_col = df_clean.columns[1]
    else:
        print(f"  * Usando primera columna '{primera_col}' como Producto")
        producto_col = primera_col

    # Seleccionar solo esas columnas y renombrar
    df_clean = df_clean[[producto_col, cantidad_col]].copy()
    df_clean.columns = ['Producto', 'Cantidad']

    # Limpiar valores nulos y vacíos
    df_clean = df_clean.dropna(subset=['Producto', 'Cantidad'])
    # Convertir Producto a string antes de usar .str accessor
    df_clean['Producto'] = df_clean['Producto'].astype(str)
    df_clean = df_clean[df_clean['Producto'].str.strip() != '']

    # Convertir cantidad a numérico (reemplazar comas por puntos primero)
    df_clean['Cantidad'] = df_clean['Cantidad'].astype(str).str.replace(',', '.')
    df_clean['Cantidad'] = pd.to_numeric(df_clean['Cantidad'], errors='coerce')
    df_clean = df_clean.dropna(subset=['Cantidad'])

    # Limpiar espacios en producto
    df_clean['Producto'] = df_clean['Producto'].str.strip()

    # Eliminar prefijos numéricos (1., 1-, 14.-, etc.) del inicio del nombre del producto
    # Patrón: número(s) + punto/guión/espacio al inicio
    df_clean['Producto'] = df_clean['Producto'].apply(
        lambda x: re.sub(r'^\d+[\.\-\s]+[\|\s]*', '', x).strip()
    )

    # Eliminar prefijos de error de OCR (I, |, i) al inicio del producto
    df_clean['Producto'] = df_clean['Producto'].apply(
        lambda x: re.sub(r'^[I\|i]\s*', '', x).strip()
    )

    return df_clean.reset_index(drop=True)


def normalizar_texto(texto: str) -> str:
    """
    Normaliza un texto eliminando espacios, puntos, guiones y caracteres especiales.
    Mantiene números y letras. Convierte todo a minúsculas para facilitar comparaciones.

    Args:
        texto: Texto a normalizar

    Returns:
        Texto normalizado (solo letras y números en minúsculas, sin espacios ni puntuación)
    """
    # Convertir a minúsculas
    texto = texto.lower().strip()

    # Eliminar puntos, guiones, espacios, y caracteres especiales comunes
    # Mantener solo letras y números
    texto_limpio = re.sub(r'[^a-záéíóúñ0-9]', '', texto)

    return texto_limpio


def validar_y_multiplicar(df_clean: pd.DataFrame, config_path: str = 'config.json') -> pd.DataFrame:
    """
    Valida los productos contra config.json y multiplica las cantidades.

    Busca coincidencias entre el producto y las variantes de entrada definidas
    en config.json usando normalización de texto (elimina espacios, puntos, números, etc.)
    para hacer comparaciones más robustas.

    Si encuentra un producto nuevo que no está en config.json, lo marca con
    "(no registrado)" y usa multiplicador 1, pero NO lo agrega al config.json.

    Soporta dos formatos de config.json:
    - Formato antiguo: {"Categoria": {"entrada": [...], "multiplicador": X}}
    - Formato nuevo: {"Categoria": {"variantes": [{"entrada": [...], "multiplicador": X}, ...]}}

    Args:
        df_clean: DataFrame limpio con Producto y Cantidad
        config_path: Ruta al archivo config.json

    Returns:
        DataFrame final con productos validados y cantidades multiplicadas
    """
    # Cargar configuración
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    resultados = []
    productos_no_encontrados = []

    for _, row in df_clean.iterrows():
        producto = row['Producto']
        cantidad = row['Cantidad']
        producto_normalizado = normalizar_texto(producto)

        encontrado = False

        # Buscar en todas las categorías
        for categoria, info in config.items():
            # Detectar formato del config.json
            if 'variantes' in info:
                # Formato nuevo: múltiples variantes con diferentes multiplicadores
                variantes = info['variantes']
            else:
                # Formato antiguo: una sola variante (retrocompatibilidad)
                variantes = [{'entrada': info['entrada'], 'multiplicador': info['multiplicador']}]

            # Buscar en cada variante
            for variante in variantes:
                entradas = variante['entrada']
                multiplicador = variante['multiplicador']

                # Verificar si el producto coincide con alguna entrada
                for entrada in entradas:
                    entrada_normalizada = normalizar_texto(entrada)

                    # Comparar textos normalizados
                    if entrada_normalizada in producto_normalizado or producto_normalizado in entrada_normalizada:
                        # Coincidencia encontrada
                        cantidad_final = cantidad * multiplicador
                        resultados.append({
                            'Producto': producto,
                            'Cantidad_Original': cantidad,
                            'Multiplicador': multiplicador,
                            'Cantidad_Final': cantidad_final,
                            'Categoria': categoria
                        })
                        encontrado = True
                        break

                if encontrado:
                    break

            if encontrado:
                break

        if not encontrado:
            # Producto sin categoría - NO agregarlo al config.json
            # Solo registrarlo con formato "(no registrado)"
            productos_no_encontrados.append(producto)

            # Crear categoría con formato especial para no registrados
            categoria_no_registrada = f"{producto} (no registrado)"

            print(f"  ! Producto no registrado: '{producto}'")

            resultados.append({
                'Producto': producto,
                'Cantidad_Original': cantidad,
                'Multiplicador': 1,
                'Cantidad_Final': cantidad,
                'Categoria': categoria_no_registrada
            })

    # Informar si hubo productos no registrados
    if productos_no_encontrados:
        print(f"\n! Se encontraron {len(productos_no_encontrados)} producto(s) no registrado(s) en config.json")
        print("   Estos aparecerán con '(no registrado)' en el Excel")

    df_final = pd.DataFrame(resultados)
    return df_final


def actualizar_inventario_layout(df_final: pd.DataFrame, layout_path: str = 'Inventario_layout.xlsx') -> None:
    """
    Actualiza el archivo Inventario_layout.xlsx con las cantidades finales por categoría.

    Preserva TODO el estilo del Excel: bordes, colores, fuentes, espaciado, etc.
    Busca la columna "entrada" en los encabezados y la fila con el nombre de la categoría,
    luego coloca el valor en la intersección.

    Args:
        df_final: DataFrame con las cantidades finales por categoría
        layout_path: Ruta al archivo Inventario_layout.xlsx
    """
    from openpyxl import load_workbook
    from copy import copy

    try:
        # Cargar el workbook existente para preservar estilos
        wb = load_workbook(layout_path)
        ws = wb.active

        # Agrupar por categoría y sumar cantidades finales
        cantidades_por_categoria = df_final.groupby('Categoria')['Cantidad_Final'].sum()

        # Buscar la columna que contiene "entrada" en la primera fila (encabezados)
        col_entrada_idx = None
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            cell_value = col[0].value
            if cell_value and str(cell_value).lower().strip() == 'entrada':
                col_entrada_idx = col_idx
                break

        if col_entrada_idx is None:
            print(f"  ! No se encontro la columna 'entrada' en {layout_path}")
            return

        # Actualizar o crear filas para cada categoría
        for categoria, cantidad in cantidades_por_categoria.items():
            if categoria == 'Sin Categoria':
                continue  # Ignorar productos sin categoría

            # Buscar la fila que contiene el nombre de la categoría en la primera columna
            fila_encontrada = False
            for fila_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=1), start=1):
                cell_value = row[0].value
                if cell_value and str(cell_value).strip() == categoria:
                    # Actualizar la celda en la intersección
                    target_cell = ws.cell(row=fila_idx, column=col_entrada_idx)
                    target_cell.value = cantidad
                    fila_encontrada = True
                    print(f"  + Actualizado '{categoria}': {cantidad}")
                    break

            if not fila_encontrada:
                # Encontrar la última fila con datos reales en la primera columna
                ultima_fila_real = 1
                for fila_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=fila_idx, column=1).value
                    if cell_value and str(cell_value).strip():
                        ultima_fila_real = fila_idx

                # Crear nueva fila inmediatamente después del último producto
                nueva_fila = ultima_fila_real + 1

                # Copiar estilos de la fila anterior si existe
                if ultima_fila_real > 1:
                    for col_idx in range(1, ws.max_column + 1):
                        celda_anterior = ws.cell(row=ultima_fila_real, column=col_idx)
                        celda_nueva = ws.cell(row=nueva_fila, column=col_idx)

                        # Copiar estilos (bordes, fuente, relleno, alineación)
                        if celda_anterior.has_style:
                            celda_nueva.font = copy(celda_anterior.font)
                            celda_nueva.border = copy(celda_anterior.border)
                            celda_nueva.fill = copy(celda_anterior.fill)
                            celda_nueva.number_format = copy(celda_anterior.number_format)
                            celda_nueva.protection = copy(celda_anterior.protection)
                            celda_nueva.alignment = copy(celda_anterior.alignment)

                # Asignar valores
                ws.cell(row=nueva_fila, column=1).value = categoria
                ws.cell(row=nueva_fila, column=col_entrada_idx).value = cantidad
                print(f"  + Creada nueva categoria '{categoria}': {cantidad}")

        # Guardar el workbook preservando todos los estilos
        wb.save(layout_path)
        print(f"\n+ Inventario actualizado exitosamente: '{layout_path}'")
        print("  (Se preservaron todos los bordes, colores y estilos del Excel)")

    except FileNotFoundError:
        print(f"\n! ADVERTENCIA: No se encontro el archivo '{layout_path}'")

    except Exception as e:
        print(f"\nX Error al actualizar inventario: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # ==================== CONFIGURACIÓN ====================
    # Comenta/Descomenta para elegir el modo de ejecución:

    # OPCIÓN 1: Extraer desde AWS Textract (usar primera vez o con nueva imagen)
    USAR_AWS = False
    #image_path = "WhatsApp Image 2025-11-04 at 10.35.29 PM (1).jpeg"
    image_path = "1 DE NOVIEMBRE 2025 (1).pdf"
    csv_path = "datos_raw.csv"

    # OPCIÓN 2: Cargar desde CSV guardado (más rápido, sin costos AWS)
    # USAR_AWS = False
    # =======================================================

    try:
        if USAR_AWS:
            # PASO 1A: Extraer tablas de la imagen con AWS
            print(f"Analizando imagen: {image_path}")
            print("Extrayendo tablas con Amazon Textract...")

            dataframes = extract_tables_from_image(image_path)

            if not dataframes:
                print("No se encontraron tablas en la imagen")
                exit(1)

            print(f"\nSe encontraron {len(dataframes)} tabla(s)")

            # Si hay múltiples tablas, filtrar y seleccionar la más grande
            # (ignorar tablas pequeñas que son encabezados)
            if len(dataframes) > 1:
                print("Filtrando tablas (ignorando encabezados)...")
                # Filtrar tablas con al menos 5 filas (probablemente son tablas de productos)
                tablas_grandes = [df for df in dataframes if len(df) >= 5]

                if tablas_grandes:
                    # Tomar la tabla más grande
                    df_raw = max(tablas_grandes, key=lambda df: len(df))
                    print(f"Tabla seleccionada: {df_raw.shape[0]} filas x {df_raw.shape[1]} columnas")
                else:
                    # Si no hay tablas grandes, tomar la más grande disponible
                    df_raw = max(dataframes, key=lambda df: len(df))
                    print(f"Tabla seleccionada (sin filtro): {df_raw.shape[0]} filas x {df_raw.shape[1]} columnas")
            else:
                df_raw = dataframes[0]

            # Guardar CSV para reutilización
            df_raw.to_csv('datos_raw.csv', index=False, encoding='utf-8-sig')
            print("DataFrame guardado en 'datos_raw.csv'")
        else:
            # PASO 1B: Cargar desde CSV
            print(f"Cargando datos desde: {csv_path}")
            df_raw = pd.read_csv(csv_path, encoding='utf-8-sig', thousands=None, decimal='.')
            print("Datos cargados exitosamente")

        # Mostrar datos raw
        print(f"\nDimensiones: {df_raw.shape[0]} filas x {df_raw.shape[1]} columnas")
        print("\nVista previa del DataFrame raw:")
        print(df_raw.head(10).to_string(index=False))

        # PASO 2: Limpiar datos (solo producto y cantidad)
        print("\n" + "="*60)
        print("PASO 2: Limpiando datos...")
        df_clean = limpiar_datos(df_raw)
        print(f"Datos limpios: {len(df_clean)} productos encontrados")
        print("\nDataFrame limpio (Producto y Cantidad):")
        print(df_clean.to_string(index=False))

        # PASO 3: Validar contra config.json y multiplicar
        print("\n" + "="*60)
        print("PASO 3: Validando productos y aplicando multiplicador...")
        df_final = validar_y_multiplicar(df_clean, 'config.json')

        if df_final.empty:
            print("\nADVERTENCIA: No se encontraron productos que coincidan con config.json")
            print("\nProductos en el pedido:")
            for producto in df_clean['Producto'].unique():
                print(f"  - {producto}")
        else:
            print(f"\nProductos validados: {len(df_final)}")
            print("\nDataFrame final:")
            print(df_final.to_string(index=False))

        # PASO 4: Exportar a Excel (siempre exportar, incluso si está vacío)
        print("\n" + "="*60)
        print("PASO 4: Exportando a Excel...")
        output_file = 'productos_final.xlsx'
        df_final.to_excel(output_file, index=False, engine='openpyxl')
        print(f"\nArchivo exportado exitosamente: '{output_file}'")

        # PASO 5: Actualizar Inventario_layout.xlsx
        if not df_final.empty:
            print("\n" + "="*60)
            print("PASO 5: Actualizando Inventario_layout.xlsx...")
            actualizar_inventario_layout(df_final, 'Inventario_layout.xlsx')

            # Resumen
            print("\n" + "="*60)
            print("RESUMEN:")
            print(f"  - Total productos validados: {len(df_final)}")
            print(f"  - Cantidad total original: {df_final['Cantidad_Original'].sum():.0f}")
            print(f"  - Cantidad total final: {df_final['Cantidad_Final'].sum():.0f}")

    except FileNotFoundError as e:
        print(f"Error: No se encontro el archivo - {str(e)}")
    except ValueError as e:
        print(f"Error de validacion: {str(e)}")
    except Exception as e:
        print(f"Error al procesar la imagen: {str(e)}")
        import traceback
        traceback.print_exc()
