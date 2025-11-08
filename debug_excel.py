from openpyxl import load_workbook

# Cargar el workbook
wb = load_workbook('Inventario_layout.xlsx')
ws = wb.active

print(f"Total de filas: {ws.max_row}")
print(f"Total de columnas: {ws.max_column}")
print("\nContenido de la primera columna (categorías):")
print("="*60)

for fila_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=1), start=1):
    cell_value = row[0].value
    if cell_value:
        print(f"Fila {fila_idx}: '{cell_value}'")
